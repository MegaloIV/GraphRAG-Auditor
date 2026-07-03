"""
Evaluación del sistema contra ground truth experto (Kappa / F1).

Módulo SEPARADO de RAGAS: RAGAS mide calidad interna sin ground truth;
aquí se compara el veredicto del sistema contra etiquetas de un experto.
Los resultados se persisten como JSON en Supabase Storage (D2):
evaluacion/{documento_id}.json.
"""
import io
import json
from datetime import datetime, timezone

import structlog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from app.core.config import get_settings
from app.models.evaluacion import (
    EvaluarRequest,
    EvaluacionResultado,
    MatrizConfusion,
    CategoriaMetricas,
)
from app.services.evaluacion.metricas_service import calcular_metricas
from app.services.grafo.neo4j_service import neo4j_service
from app.services.storage.supabase_storage_service import storage_service

logger = structlog.get_logger(__name__)
settings = get_settings()
router = APIRouter(prefix="/evaluacion", tags=["Evaluación (ground truth)"])


def _objeto_resultado(documento_id: str) -> str:
    return f"evaluacion/{documento_id}.json"


def _leer_veredictos_sistema(documento_id: str) -> dict[str, str]:
    """{cita_id: veredicto} de las citas ya auditadas del documento."""
    query = """
    MATCH (d:Documento {id: $doc_id})-[:TIENE_CITA]->(c:Cita)
    WHERE c.veredicto IS NOT NULL
    RETURN c.id AS cita_id, c.veredicto AS veredicto
    """
    with neo4j_service.driver.session(database=settings.neo4j_database) as session:
        return {
            rec["cita_id"]: rec["veredicto"]
            for rec in session.run(query, doc_id=documento_id)
        }


@router.post(
    "/{documento_id}/evaluar",
    response_model=EvaluacionResultado,
    summary="Evaluar los veredictos del sistema contra etiquetas de un experto",
)
async def evaluar(documento_id: str, solicitud: EvaluarRequest):
    """
    Empareja por cita_id los veredictos del sistema (Neo4j) con las etiquetas
    del experto recibidas en el body (D3), calcula matriz de confusión,
    P/R/F1 (categoría + macro + weighted) y Kappa de Cohen, persiste el
    resultado en Storage y lo retorna. Las citas sin contraparte se reportan
    como no emparejadas, sin romper el cálculo.
    """
    # TODO(B3): sustituir el body por ground_truth_loader.cargar(documento_id)
    # cuando se defina el mecanismo de carga del ground truth.
    try:
        veredictos = _leer_veredictos_sistema(documento_id)
    except Exception as e:
        logger.error("error_leer_veredictos", doc_id=documento_id, error=str(e))
        raise HTTPException(status_code=500, detail={
            "codigo": "ERROR_INTERNO",
            "mensaje": str(e),
            "accion_sugerida": "Intenta nuevamente.",
        })

    if not veredictos:
        raise HTTPException(status_code=404, detail={
            "codigo": "SIN_VEREDICTOS",
            "mensaje": "El documento no tiene citas auditadas.",
            "accion_sugerida": "Ejecuta primero POST /auditoria/{documento_id}/auditar.",
        })

    etiquetas = {e.cita_id: e.etiqueta_experto.value for e in solicitud.etiquetas}

    pares = [
        (veredictos[cita_id], etiqueta)
        for cita_id, etiqueta in etiquetas.items()
        if cita_id in veredictos
    ]
    no_emparejadas_sistema = sum(1 for cid in veredictos if cid not in etiquetas)
    no_emparejadas_experto = sum(1 for cid in etiquetas if cid not in veredictos)

    if not pares:
        raise HTTPException(status_code=400, detail={
            "codigo": "SIN_EMPAREJAMIENTOS",
            "mensaje": "Ninguna etiqueta del experto coincide con una cita auditada.",
            "accion_sugerida": "Verifica que los cita_id correspondan al documento.",
        })

    metricas = calcular_metricas(pares)

    resultado = EvaluacionResultado(
        documento_id=documento_id,
        total_evaluadas=metricas["total_evaluadas"],
        aciertos=metricas["aciertos"],
        kappa_cohen=metricas["kappa_cohen"],
        matriz_confusion=MatrizConfusion(
            labels=metricas["labels"],
            filas=metricas["matriz"],
        ),
        por_categoria=[CategoriaMetricas(**c) for c in metricas["por_categoria"]],
        macro=CategoriaMetricas(**metricas["macro"]),
        weighted=CategoriaMetricas(**metricas["weighted"]),
        no_emparejadas_sistema=no_emparejadas_sistema,
        no_emparejadas_experto=no_emparejadas_experto,
        evaluado_en=datetime.now(timezone.utc).isoformat(),
    )

    # D2: persistir como JSON en Supabase Storage.
    storage_service.subir_texto(
        _objeto_resultado(documento_id), resultado.model_dump_json()
    )
    logger.info(
        "evaluacion_completada",
        doc_id=documento_id,
        total=resultado.total_evaluadas,
        kappa=resultado.kappa_cohen,
    )
    return resultado


def _cargar_resultado(documento_id: str) -> EvaluacionResultado:
    crudo = storage_service.leer_texto(_objeto_resultado(documento_id))
    if crudo is None:
        raise HTTPException(status_code=404, detail={
            "codigo": "EVALUACION_NO_ENCONTRADA",
            "mensaje": "No hay una evaluación persistida para este documento.",
            "accion_sugerida": "Ejecuta primero POST /evaluacion/{documento_id}/evaluar.",
        })
    return EvaluacionResultado(**json.loads(crudo))


@router.get(
    "/{documento_id}/resultados",
    response_model=EvaluacionResultado,
    summary="Última evaluación persistida del documento",
)
async def ver_resultados(documento_id: str):
    return _cargar_resultado(documento_id)


@router.get(
    "/{documento_id}/exportar",
    summary="Exportar la evaluación a Excel (hojas Resumen y Matriz)",
)
async def exportar_evaluacion(documento_id: str):
    resultado = _cargar_resultado(documento_id)

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Métrica", "Valor"])
    for celda in ws[1]:
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = center

    filas_resumen = [
        ("Documento", resultado.documento_id),
        ("Evaluado en", resultado.evaluado_en),
        ("Total evaluadas", resultado.total_evaluadas),
        ("Aciertos", resultado.aciertos),
        ("Kappa de Cohen", resultado.kappa_cohen),
        ("No emparejadas (sistema)", resultado.no_emparejadas_sistema),
        ("No emparejadas (experto)", resultado.no_emparejadas_experto),
    ]
    for etiqueta, valor in filas_resumen:
        ws.append([etiqueta, valor])

    ws.append([])
    ws.append(["Categoría", "Precisión", "Recall", "F1", "Soporte"])
    for celda in ws[ws.max_row]:
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = center
    for cat in [*resultado.por_categoria, resultado.macro, resultado.weighted]:
        ws.append([cat.categoria, cat.precision, cat.recall, cat.f1, cat.soporte])

    ws.column_dimensions["A"].width = 28
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 14

    # ── Hoja 2: Matriz de confusión ──────────────────────────────────────────
    ws_m = wb.create_sheet("Matriz")
    labels = resultado.matriz_confusion.labels
    ws_m.append(["Experto \\ Sistema", *labels])
    for celda in ws_m[1]:
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = center
    for label, fila in zip(labels, resultado.matriz_confusion.filas):
        ws_m.append([label, *fila])
        ws_m.cell(row=ws_m.max_row, column=1).font = Font(bold=True)
    ws_m.column_dimensions["A"].width = 20
    for col in ("B", "C", "D"):
        ws_m.column_dimensions[col].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info("exportar_evaluacion_excel", doc_id=documento_id)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="evaluacion_{documento_id}.xlsx"'
        },
    )
