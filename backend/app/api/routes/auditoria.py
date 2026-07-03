"""
EP-004: Endpoints de Auditoría Semántica

HU-010  POST /{doc_id}/auditar              → auditar todas las citas del documento
HU-010  GET  /{doc_id}/veredictos           → ver veredictos ya calculados
HU-011  GET  /{doc_id}/alertas              → inconsistencias estructurales
HU-012  GET  /{doc_id}/alertas/alucinaciones → citas que el sistema no pudo verificar
"""
import io
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import structlog

from app.core.config import get_settings
from app.services.auditoria.auditoria_service import auditoria_service
from app.services.grafo.neo4j_service import neo4j_service
from app.models.auditoria import (
    VeredictoAuditoria,
    VeredictoTipo,
    AuditoriaResponse,
    AlertasResponse,
    AlertaInconsistencia,
    AlertasAlucinacionResponse,
    MetricasRagasResponse,
)

logger = structlog.get_logger(__name__)
settings = get_settings()
router = APIRouter(prefix="/auditoria", tags=["Auditoría"])


# ── HU-010: Auditar documento ────────────────────────────────────────────────

@router.post(
    "/{documento_id}/auditar",
    response_model=AuditoriaResponse,
    summary="HU-010: Auditar todas las citas del documento",
)
async def auditar_documento(documento_id: str):
    """
    Ejecuta la auditoría semántica completa:
    por cada cita recupera evidencia del grafo + Supabase/pgvector
    y emite un veredicto con gpt-5.4-mini.
    Persiste los veredictos en Neo4j.
    """
    try:
        veredictos = auditoria_service.auditar_documento(documento_id)

        if not veredictos:
            raise HTTPException(
                status_code=404,
                detail={
                    "codigo": "SIN_CITAS",
                    "mensaje": "No se encontraron citas para auditar en este documento.",
                    "accion_sugerida": "Verifica que la auditoría inicial se haya completado.",
                },
            )

        supports    = sum(1 for v in veredictos if v.veredicto == VeredictoTipo.SUPPORTS)
        refutes     = sum(1 for v in veredictos if v.veredicto == VeredictoTipo.REFUTES)
        no_info     = sum(1 for v in veredictos if v.veredicto == VeredictoTipo.NO_INFO)

        advertencia = None
        if no_info > 0:
            advertencia = (
                f"{no_info} cita(s) no pudieron verificarse por falta de evidencia. "
                f"Revisa las alertas de verificación."
            )

        return AuditoriaResponse(
            documento_id=documento_id,
            total_citas=len(veredictos),
            supports=supports,
            refutes=refutes,
            no_info=no_info,
            veredictos=veredictos,
            advertencia=advertencia,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error("error_auditar_documento", doc_id=documento_id, error=str(e))
        raise HTTPException(
            status_code=500,
            detail={"codigo": "ERROR_INTERNO", "mensaje": str(e),
                    "accion_sugerida": "Intenta nuevamente."},
        )


# ── HU-010: Ver veredictos ya calculados ─────────────────────────────────────

@router.get(
    "/{documento_id}/veredictos",
    response_model=AuditoriaResponse,
    summary="HU-010: Ver veredictos de auditoría ya calculados",
)
async def ver_veredictos(documento_id: str):
    """
    Lee los veredictos persistidos en Neo4j sin volver a llamar al LLM.
    """
    query = """
    MATCH (d:Documento {id: $doc_id})-[:TIENE_CITA]->(c:Cita)
    WHERE c.veredicto IS NOT NULL
    OPTIONAL MATCH (c)-[:CITA_A]->(r:Referencia)
    OPTIONAL MATCH (r)-[:ESCRITO_POR]->(a:Autor)
    RETURN
      c.id            AS cita_id,
      c.texto         AS texto_cita,
      c.fragmento     AS fragmento_oracion,
      c.pagina        AS pagina,
      c.veredicto     AS veredicto,
      c.justificacion AS justificacion,
      c.pagina_paper  AS pagina_paper,
      c.faithfulness      AS faithfulness,
      c.answer_relevancy  AS answer_relevancy,
      c.context_precision AS context_precision,
      r.id            AS ref_id,
      r.titulo_oficial AS ref_titulo_oficial,
      r.titulo        AS ref_titulo,
      r.anio          AS ref_anio,
      r.doi_verificado AS doi,
      collect(DISTINCT a.nombre) AS autores
    """
    try:
        with neo4j_service.driver.session(database=settings.neo4j_database) as session:
            registros = list(session.run(query, doc_id=documento_id))

        if not registros:
            raise HTTPException(
                status_code=404,
                detail={
                    "codigo": "VEREDICTOS_NO_ENCONTRADOS",
                    "mensaje": "No hay veredictos calculados para este documento.",
                    "accion_sugerida": "Ejecuta primero POST /auditoria/{documento_id}/auditar.",
                },
            )

        veredictos = []
        for r in registros:
            try:
                tipo = VeredictoTipo(r["veredicto"])
            except ValueError:
                tipo = VeredictoTipo.NO_INFO

            veredictos.append(VeredictoAuditoria(
                cita_id=r["cita_id"],
                texto_cita=r["texto_cita"] or "",
                fragmento_oracion=r["fragmento_oracion"] or "",
                pagina=r["pagina"] or 0,
                veredicto=tipo,
                justificacion=r["justificacion"] or "",
                referencia_id=r["ref_id"],
                titulo_referencia=r["ref_titulo_oficial"] or r["ref_titulo"],
                anio_referencia=r["ref_anio"],
                doi_referencia=r["doi"],
                autores_referencia=r["autores"] or [],
                pagina_paper=r.get("pagina_paper"),
                faithfulness=r.get("faithfulness"),
                answer_relevancy=r.get("answer_relevancy"),
                context_precision=r.get("context_precision"),
            ))

        supports = sum(1 for v in veredictos if v.veredicto == VeredictoTipo.SUPPORTS)
        refutes  = sum(1 for v in veredictos if v.veredicto == VeredictoTipo.REFUTES)
        no_info  = sum(1 for v in veredictos if v.veredicto == VeredictoTipo.NO_INFO)

        return AuditoriaResponse(
            documento_id=documento_id,
            total_citas=len(veredictos),
            supports=supports,
            refutes=refutes,
            no_info=no_info,
            veredictos=veredictos,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error("error_ver_veredictos", doc_id=documento_id, error=str(e))
        raise HTTPException(
            status_code=500,
            detail={"codigo": "ERROR_INTERNO", "mensaje": str(e),
                    "accion_sugerida": "Intenta nuevamente."},
        )


# ── HU-011: Alertas de inconsistencias estructurales ────────────────────────

@router.get(
    "/{documento_id}/alertas",
    response_model=AlertasResponse,
    summary="HU-011: Ver alertas de citas sin referencia y referencias sin citar",
)
async def ver_alertas(documento_id: str):
    """
    Detecta inconsistencias estructurales:
    - Citas en el cuerpo sin referencia bibliográfica correspondiente
    - Referencias listadas que nunca se citan en el texto
    """
    try:
        resultado = auditoria_service.detectar_inconsistencias(documento_id)

        citas_sin_ref   = resultado["citas_sin_referencia"]
        refs_sin_citar  = resultado["referencias_sin_citar"]
        total           = len(citas_sin_ref) + len(refs_sin_citar)

        if total == 0:
            mensaje = "No se detectaron inconsistencias estructurales. Todas las citas tienen referencia y todas las referencias están citadas."
        else:
            partes = []
            if citas_sin_ref:
                partes.append(f"{len(citas_sin_ref)} cita(s) sin referencia")
            if refs_sin_citar:
                partes.append(f"{len(refs_sin_citar)} referencia(s) sin citar")
            mensaje = "Se detectaron inconsistencias: " + " y ".join(partes) + "."

        return AlertasResponse(
            documento_id=documento_id,
            citas_sin_referencia=citas_sin_ref,
            referencias_sin_citar=refs_sin_citar,
            total_inconsistencias=total,
            mensaje=mensaje,
        )

    except Exception as e:
        logger.error("error_ver_alertas", doc_id=documento_id, error=str(e))
        raise HTTPException(
            status_code=500,
            detail={"codigo": "ERROR_INTERNO", "mensaje": str(e),
                    "accion_sugerida": "Intenta nuevamente."},
        )


# ── HU-012: Alertas de alucinaciones del sistema ────────────────────────────

@router.get(
    "/{documento_id}/alertas/alucinaciones",
    response_model=AlertasAlucinacionResponse,
    summary="HU-012: Ver alertas de citas que el sistema no pudo verificar",
)
async def ver_alertas_alucinaciones(documento_id: str):
    """
    Lista las citas con veredicto NO_VERIFICABLE — aquellas para las que
    el sistema no encontró evidencia en el grafo ni en Supabase/pgvector.
    El revisor no debe confiar ciegamente en estos casos.
    """
    query = """
    MATCH (d:Documento {id: $doc_id})-[:TIENE_CITA]->(c:Cita)
    WHERE c.veredicto = $veredicto
    RETURN c.id AS cita_id, c.texto AS texto_cita,
           c.pagina AS pagina, c.justificacion AS justificacion
    """
    try:
        with neo4j_service.driver.session(database=settings.neo4j_database) as session:
            registros = list(session.run(
                query,
                doc_id=documento_id,
                veredicto=VeredictoTipo.NO_INFO.value,
            ))

        from app.models.auditoria import AlertaAlucinacionSistema
        alertas = [
            AlertaAlucinacionSistema(
                cita_id=r["cita_id"],
                texto_cita=r["texto_cita"] or "",
                pagina=r["pagina"] or 0,
                razon_no_verificable=r["justificacion"] or "",
            )
            for r in registros
        ]

        advertencia = None
        if alertas:
            advertencia = (
                f"Se encontraron {len(alertas)} cita(s) que el sistema no pudo verificar. "
                f"Estas citas requieren revisión manual."
            )

        return AlertasAlucinacionResponse(
            documento_id=documento_id,
            total_no_info=len(alertas),
            alertas=alertas,
            advertencia=advertencia,
        )

    except Exception as e:
        logger.error("error_alertas_alucinaciones", doc_id=documento_id, error=str(e))
        raise HTTPException(
            status_code=500,
            detail={"codigo": "ERROR_INTERNO", "mensaje": str(e),
                    "accion_sugerida": "Intenta nuevamente."},
        )


# ── EP-RAGAS: Evaluación RAGAS del documento ────────────────────────────────

@router.post(
    "/{documento_id}/evaluar-ragas",
    response_model=MetricasRagasResponse,
    summary="EP-RAGAS: Evaluar las citas con fragmento de paper disponible",
)
async def evaluar_ragas(documento_id: str):
    """
    Evalúa con RAGAS las citas que tienen fragmento de evidencia.
    Persiste los scores en cada Cita y retorna los promedios.
    """
    try:
        resultado = auditoria_service.evaluar_ragas_documento(documento_id)

        if resultado.get("total_evaluadas", 0) == 0:
            raise HTTPException(
                status_code=404,
                detail={
                    "codigo": "SIN_EVIDENCIA",
                    "mensaje": (
                        "No hay citas con fragmento de paper disponible "
                        "para evaluar. Verifica que los papers estén indexados."
                    ),
                    "accion_sugerida": "Indexa los papers citados y vuelve a auditar.",
                },
            )

        return MetricasRagasResponse(**resultado)

    except HTTPException:
        raise
    except Exception as e:
        logger.error("error_evaluar_ragas", doc_id=documento_id, error=str(e))
        raise HTTPException(
            status_code=500,
            detail={"codigo": "ERROR_INTERNO", "mensaje": str(e),
                    "accion_sugerida": "Intenta nuevamente."},
        )


@router.get(
    "/{documento_id}/metricas",
    response_model=MetricasRagasResponse,
    summary="EP-RAGAS: Ver métricas RAGAS promedio del documento",
)
async def ver_metricas(documento_id: str):
    """
    Lee los scores RAGAS ya calculados en cada Cita y retorna
    los promedios del documento.
    """
    query = """
    MATCH (d:Documento {id: $doc_id})-[:TIENE_CITA]->(c:Cita)
    WHERE c.faithfulness IS NOT NULL
    RETURN
      avg(c.faithfulness)      AS faithfulness_promedio,
      avg(c.answer_relevancy)  AS answer_relevancy_promedio,
      avg(c.context_precision) AS context_precision_promedio,
      count(c)                 AS total_citas_evaluadas
    """
    try:
        with neo4j_service.driver.session(database=settings.neo4j_database) as session:
            registro = session.run(query, doc_id=documento_id).single()

        total = registro["total_citas_evaluadas"] if registro else 0
        if total == 0:
            raise HTTPException(
                status_code=404,
                detail={
                    "codigo": "METRICAS_NO_ENCONTRADAS",
                    "mensaje": "No hay métricas RAGAS calculadas para este documento.",
                    "accion_sugerida": "Ejecuta primero POST /auditoria/{documento_id}/evaluar-ragas.",
                },
            )

        def _redondear(v):
            return round(float(v), 3) if v is not None else None

        return MetricasRagasResponse(
            total_evaluadas=total,
            faithfulness_promedio=_redondear(registro["faithfulness_promedio"]),
            answer_relevancy_promedio=_redondear(registro["answer_relevancy_promedio"]),
            context_precision_promedio=_redondear(registro["context_precision_promedio"]),
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error("error_ver_metricas", doc_id=documento_id, error=str(e))
        raise HTTPException(
            status_code=500,
            detail={"codigo": "ERROR_INTERNO", "mensaje": str(e),
                    "accion_sugerida": "Intenta nuevamente."},
        )


# ── Exportar informe completo a Excel ───────────────────────────────────────

@router.get(
    "/{documento_id}/metricas/exportar",
    summary="Exportar informe completo de auditoría a Excel",
)
async def exportar_metricas_excel(documento_id: str):
    """
    Genera un Excel con dos hojas:
    - Resumen: estadísticas generales del documento
    - Citas:   detalle por cita con veredicto, fragmentos y métricas RAGAS
    """
    _Q_CITAS_DETALLE = """
    MATCH (d:Documento {id: $doc_id})-[:TIENE_CITA]->(c:Cita)
    OPTIONAL MATCH (c)-[:CITA_A]->(r:Referencia)
    OPTIONAL MATCH (r)-[:ESCRITO_POR]->(a:Autor)
    RETURN
      c.texto              AS texto_cita,
      c.fragmento          AS fragmento_oracion,
      c.pagina             AS pagina_tesis,
      c.veredicto          AS veredicto,
      c.justificacion      AS justificacion,
      c.fragmento_evidencia AS fragmento_paper,
      c.pagina_paper       AS pagina_paper,
      c.faithfulness       AS faithfulness,
      c.answer_relevancy   AS answer_relevancy,
      c.context_precision  AS context_precision,
      r.titulo_oficial     AS titulo_oficial,
      r.titulo             AS titulo_referencia,
      r.anio               AS anio_referencia,
      collect(DISTINCT a.nombre) AS autores
    ORDER BY c.pagina
    """

    _Q_REFERENCIAS = """
    MATCH (d:Documento {id: $doc_id})-[:TIENE_REFERENCIA]->(r:Referencia)
    RETURN count(r) AS total_referencias,
           sum(CASE WHEN r.nivel_confianza IS NOT NULL AND r.nivel_confianza <> 'no_encontrado' THEN 1 ELSE 0 END) AS con_texto
    """

    _Q_REFS_SIN_CITAR = """
    MATCH (d:Documento {id: $doc_id})-[:TIENE_REFERENCIA]->(r:Referencia)
    WHERE NOT (r)<-[:CITA_A]-()
    RETURN count(r) AS total
    """

    _Q_CITAS_SIN_REF = """
    MATCH (d:Documento {id: $doc_id})-[:TIENE_CITA]->(c:Cita)
    WHERE NOT (c)-[:CITA_A]->()
    RETURN count(c) AS total
    """

    _Q_RAGAS_PROMEDIO = """
    MATCH (d:Documento {id: $doc_id})-[:TIENE_CITA]->(c:Cita)
    WHERE c.faithfulness IS NOT NULL
    RETURN
      count(c)               AS total_evaluadas,
      avg(c.faithfulness)    AS faithfulness_avg,
      avg(c.answer_relevancy) AS relevancy_avg,
      avg(c.context_precision) AS precision_avg
    """

    try:
        with neo4j_service.driver.session(database=settings.neo4j_database) as session:
            citas_raw   = list(session.run(_Q_CITAS_DETALLE, doc_id=documento_id))
            refs_raw    = session.run(_Q_REFERENCIAS, doc_id=documento_id).single()
            sin_citar   = session.run(_Q_REFS_SIN_CITAR, doc_id=documento_id).single()
            sin_ref     = session.run(_Q_CITAS_SIN_REF, doc_id=documento_id).single()
            ragas_prom  = session.run(_Q_RAGAS_PROMEDIO, doc_id=documento_id).single()

        if not citas_raw:
            raise HTTPException(
                status_code=404,
                detail={
                    "codigo": "CITAS_NO_ENCONTRADAS",
                    "mensaje": "No hay citas auditadas para exportar.",
                    "accion_sugerida": "Ejecuta primero POST /auditoria/{documento_id}/auditar.",
                },
            )

        def _r(v):
            return round(float(v), 3) if v is not None else None

        # ── Estilos compartidos ──────────────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        section_fill = PatternFill("solid", fgColor="334155")
        section_font = Font(bold=True, color="CBD5E1", size=10)
        center = Alignment(horizontal="center", vertical="center")
        wrap   = Alignment(wrap_text=True, vertical="top")

        FILL_SUPPORTS = PatternFill("solid", fgColor="D1FAE5")
        FILL_REFUTES  = PatternFill("solid", fgColor="FEE2E2")
        FILL_NO_INFO  = PatternFill("solid", fgColor="F1F5F9")

        wb = openpyxl.Workbook()

        # ── Hoja 1: Resumen ──────────────────────────────────────────────────
        ws_res = wb.active
        ws_res.title = "Resumen"

        def _escribir_seccion(ws, fila, titulo):
            c = ws.cell(row=fila, column=1, value=titulo)
            c.fill = section_fill
            c.font = section_font
            c.alignment = center
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
            ws.row_dimensions[fila].height = 20
            return fila + 1

        def _escribir_fila(ws, fila, etiqueta, valor):
            ws.cell(row=fila, column=1, value=etiqueta).font = Font(bold=False, size=10)
            ws.cell(row=fila, column=2, value=valor).alignment = center
            return fila + 1

        fila = 1
        fila = _escribir_seccion(ws_res, fila, "REFERENCIAS")
        total_refs = refs_raw["total_referencias"] if refs_raw else 0
        con_texto  = refs_raw["con_texto"] if refs_raw else 0
        fila = _escribir_fila(ws_res, fila, "Total referencias bibliográficas", total_refs)
        fila = _escribir_fila(ws_res, fila, "Referencias con texto disponible", con_texto)
        fila = _escribir_fila(ws_res, fila, "Referencias no encontradas", total_refs - con_texto)
        fila = _escribir_fila(ws_res, fila, "Referencias sin citar en el texto", sin_citar["total"] if sin_citar else 0)

        fila += 1
        fila = _escribir_seccion(ws_res, fila, "CITAS")
        total_citas   = len(citas_raw)
        supports_cnt  = sum(1 for r in citas_raw if r["veredicto"] == "SUPPORTS")
        refutes_cnt   = sum(1 for r in citas_raw if r["veredicto"] == "REFUTES")
        no_info_cnt   = sum(1 for r in citas_raw if r["veredicto"] == "NO_INFO")
        sin_ref_cnt   = sin_ref["total"] if sin_ref else 0
        fila = _escribir_fila(ws_res, fila, "Total citas identificadas", total_citas)
        fila = _escribir_fila(ws_res, fila, "Citas sin referencia bibliográfica", sin_ref_cnt)
        fila = _escribir_fila(ws_res, fila, "Veredicto SUPPORTS (respaldadas)", supports_cnt)
        fila = _escribir_fila(ws_res, fila, "Veredicto REFUTES (refutadas)", refutes_cnt)
        fila = _escribir_fila(ws_res, fila, "Veredicto NO_INFO (no verificables)", no_info_cnt)

        fila += 1
        total_ev = ragas_prom["total_evaluadas"] if ragas_prom else 0
        if total_ev:
            fila = _escribir_seccion(ws_res, fila, "MÉTRICAS RAGAS")
            fila = _escribir_fila(ws_res, fila, "Citas evaluadas con RAGAS", int(total_ev))
            fila = _escribir_fila(ws_res, fila, "Faithfulness (promedio)", _r(ragas_prom["faithfulness_avg"]))
            fila = _escribir_fila(ws_res, fila, "Answer Relevancy (promedio)", _r(ragas_prom["relevancy_avg"]))
            fila = _escribir_fila(ws_res, fila, "Context Precision (promedio)", _r(ragas_prom["precision_avg"]))

        ws_res.column_dimensions["A"].width = 42
        ws_res.column_dimensions["B"].width = 16

        # ── Hoja 2: Citas ────────────────────────────────────────────────────
        ws_citas = wb.create_sheet("Citas")

        encabezados = [
            "Paper (título)",
            "Cita APA",
            "Afirmación del tesista",
            "Fragmento del paper",
            "Pág. tesis",
            "Pág. paper",
            "Veredicto",
            "Justificación",
            "Faithfulness",
            "Answer Relevancy",
            "Context Precision",
        ]

        for col_idx, titulo in enumerate(encabezados, start=1):
            c = ws_citas.cell(row=1, column=col_idx, value=titulo)
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_align
        ws_citas.row_dimensions[1].height = 32

        for fila_idx, reg in enumerate(citas_raw, start=2):
            titulo_paper = reg["titulo_oficial"] or reg["titulo_referencia"] or ""
            ws_citas.cell(row=fila_idx, column=1,  value=titulo_paper).alignment = wrap
            ws_citas.cell(row=fila_idx, column=2,  value=reg["texto_cita"] or "").alignment = wrap
            ws_citas.cell(row=fila_idx, column=3,  value=reg["fragmento_oracion"] or "").alignment = wrap
            ws_citas.cell(row=fila_idx, column=4,  value=reg["fragmento_paper"] or "").alignment = wrap
            ws_citas.cell(row=fila_idx, column=5,  value=reg["pagina_tesis"] or "").alignment = center
            ws_citas.cell(row=fila_idx, column=6,  value=reg["pagina_paper"] or "").alignment = center
            ws_citas.cell(row=fila_idx, column=7,  value=reg["veredicto"] or "").alignment = center
            ws_citas.cell(row=fila_idx, column=8,  value=reg["justificacion"] or "").alignment = wrap
            ws_citas.cell(row=fila_idx, column=9,  value=_r(reg["faithfulness"])).alignment = center
            ws_citas.cell(row=fila_idx, column=10, value=_r(reg["answer_relevancy"])).alignment = center
            ws_citas.cell(row=fila_idx, column=11, value=_r(reg["context_precision"])).alignment = center

            veredicto = reg["veredicto"] or ""
            if veredicto == "SUPPORTS":
                row_fill = FILL_SUPPORTS
            elif veredicto == "REFUTES":
                row_fill = FILL_REFUTES
            elif veredicto == "NO_INFO":
                row_fill = FILL_NO_INFO
            else:
                row_fill = None

            if row_fill:
                for col_idx in range(1, 12):
                    ws_citas.cell(row=fila_idx, column=col_idx).fill = row_fill

        ws_citas.column_dimensions["A"].width = 40
        ws_citas.column_dimensions["B"].width = 22
        ws_citas.column_dimensions["C"].width = 45
        ws_citas.column_dimensions["D"].width = 55
        ws_citas.column_dimensions["E"].width = 10
        ws_citas.column_dimensions["F"].width = 10
        ws_citas.column_dimensions["G"].width = 14
        ws_citas.column_dimensions["H"].width = 45
        ws_citas.column_dimensions["I"].width = 15
        ws_citas.column_dimensions["J"].width = 17
        ws_citas.column_dimensions["K"].width = 17
        ws_citas.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"informe_{documento_id}.xlsx"
        headers = {"Content-Disposition": f"attachment; filename={filename}"}

        logger.info("exportar_informe_excel", doc_id=documento_id, total_citas=len(citas_raw))
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error("error_exportar_informe", doc_id=documento_id, error=str(e))
        raise HTTPException(
            status_code=500,
            detail={"codigo": "ERROR_INTERNO", "mensaje": str(e),
                    "accion_sugerida": "Intenta nuevamente."},
        )